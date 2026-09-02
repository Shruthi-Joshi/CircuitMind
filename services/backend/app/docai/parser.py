"""BOM file parser — extracts structured line items from PDF, XLSX, CSV, and
plain-text uploads.

Each parser returns a list of ``BOMRow`` dicts ready for database insertion.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

# PDF
try:
    from pypdf import PdfReader  # type: ignore
except ImportError:  # pragma: no cover
    PdfReader = None  # type: ignore

# XLSX
try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore


@dataclass
class BOMRow:
    """Normalised BOM line item extracted from any file format."""
    line_number: int
    reference_designator: str
    mpn: str
    quantity: int
    description: str


# ── Heuristic MPN detector ───────────────────────────────────────────────────

_MPN_PATTERN = re.compile(
    r"[A-Z]{2,}[\dA-Z]*[-/][\dA-Za-z.]{3,}",
    re.IGNORECASE,
)

_QTY_PATTERN = re.compile(r"\b(\d{1,6})\b")

_REF_DES_PATTERN = re.compile(
    r"\b([CRULDJTQSPV]\d{1,4})\b",  # C1, R12, U3, L1, D5, etc.
    re.IGNORECASE,
)


def _try_int(val: Any) -> int:
    """Best-effort integer cast, default 1."""
    try:
        return max(1, int(float(val)))
    except (TypeError, ValueError):
        return 1


def _clean(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


# ── Format-specific extractors ────────────────────────────────────────────────

def parse_pdf(filepath: str | Path) -> list[dict]:
    """Extract BOM rows from a PDF using pypdf text extraction."""
    if PdfReader is None:
        raise RuntimeError("pypdf not installed")
    reader = PdfReader(str(filepath))
    full_text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return _parse_text(full_text)


def parse_xlsx(filepath: str | Path) -> list[dict]:
    """Extract BOM rows from an Excel workbook (first sheet)."""
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_raw: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append([_clean(c) for c in row])
    wb.close()
    return _parse_tabular(rows_raw)


def parse_csv(filepath: str | Path) -> list[dict]:
    """Extract BOM rows from a CSV / TSV file."""
    text = Path(filepath).read_text(encoding="utf-8", errors="replace")
    dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    reader = csv.reader(io.StringIO(text), dialect)
    rows_raw = [[_clean(c) for c in row] for row in reader]
    return _parse_tabular(rows_raw)


def parse_text(text: str) -> list[dict]:
    """Parse plain text (e.g. pasted BOM list) into rows."""
    return _parse_text(text)


# ── Core parsing helpers ─────────────────────────────────────────────────────

def _guess_columns(header: list[str]) -> dict[str, int | None]:
    """Map header labels to logical columns via keyword matching."""
    mapping: dict[str, int | None] = {
        "mpn": None,
        "ref": None,
        "qty": None,
        "desc": None,
    }
    kw = {
        "mpn": ["mpn", "part", "mfr part", "manufacturer part", "mfg part", "p/n", "part number", "part no"],
        "ref": ["ref", "designator", "reference", "refdes"],
        "qty": ["qty", "quantity", "count", "amount"],
        "desc": ["desc", "description", "name", "component", "value"],
    }
    for idx, cell in enumerate(header):
        low = cell.lower()
        for key, keywords in kw.items():
            if any(k in low for k in keywords):
                mapping[key] = idx
    return mapping


def _parse_tabular(rows: list[list[str]]) -> list[dict]:
    """Shared parser for XLSX / CSV where data comes in row-of-cells format."""
    if not rows:
        return []

    # Use first non-empty row as header
    header_idx = 0
    for i, r in enumerate(rows):
        if any(r):
            header_idx = i
            break

    col_map = _guess_columns(rows[header_idx])
    results: list[dict] = []
    line_num = 0

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue

        mpn = ""
        ref = ""
        qty = 1
        desc = ""

        if col_map["mpn"] is not None and col_map["mpn"] < len(row):
            mpn = row[col_map["mpn"]]
        if col_map["ref"] is not None and col_map["ref"] < len(row):
            ref = row[col_map["ref"]]
        if col_map["qty"] is not None and col_map["qty"] < len(row):
            qty = _try_int(row[col_map["qty"]])
        if col_map["desc"] is not None and col_map["desc"] < len(row):
            desc = row[col_map["desc"]]

        # If no MPN column identified, try to detect MPN in any cell
        if not mpn:
            for cell in row:
                m = _MPN_PATTERN.search(cell)
                if m:
                    mpn = m.group()
                    break

        if not mpn:
            continue

        # Fallback: grab ref designator from any cell
        if not ref:
            for cell in row:
                m = _REF_DES_PATTERN.search(cell)
                if m:
                    ref = m.group()
                    break

        # Fallback description = all cells joined
        if not desc:
            desc = " ".join(c for c in row if c and c != mpn and c != ref)

        line_num += 1
        results.append(asdict(BOMRow(
            line_number=line_num,
            reference_designator=ref,
            mpn=mpn.strip(),
            quantity=qty,
            description=desc.strip(),
        )))

    return results


def _parse_text(text: str) -> list[dict]:
    """Parse unstructured text (from PDF or pasted content) line-by-line."""
    results: list[dict] = []
    line_num = 0

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        
        # Skip header lines and decorative elements
        if any(skip in line.lower() for skip in ['bom', 'reference', 'part number', 'qty', 'description', '---', '===', 'rev ', 'total:', 'notes:']):
            continue
        
        # Handle pipe-separated table format (like our power supply BOM)
        if '|' in line:
            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 4:  # ref | mpn | qty | desc
                ref = parts[0].strip()
                mpn_candidate = parts[1].strip()
                qty_candidate = parts[2].strip()
                desc = parts[3].strip() if len(parts) > 3 else ""
                
                # Validate MPN
                if _MPN_PATTERN.search(mpn_candidate):
                    line_num += 1
                    results.append(asdict(BOMRow(
                        line_number=line_num,
                        reference_designator=ref,
                        mpn=mpn_candidate,
                        quantity=_try_int(qty_candidate),
                        description=desc,
                    )))
                    continue

        # Original MPN detection logic for other formats
        mpn_match = _MPN_PATTERN.search(line)
        if not mpn_match:
            continue

        mpn = mpn_match.group()
        ref = ""
        qty = 1

        ref_match = _REF_DES_PATTERN.search(line)
        if ref_match:
            ref = ref_match.group()

        # Look for a standalone number as quantity
        remaining = line.replace(mpn, "").replace(ref, "")
        qty_match = _QTY_PATTERN.search(remaining)
        if qty_match:
            qty = _try_int(qty_match.group(1))

        desc = remaining.strip(" \t,;|")

        line_num += 1
        results.append(asdict(BOMRow(
            line_number=line_num,
            reference_designator=ref,
            mpn=mpn.strip(),
            quantity=qty,
            description=desc.strip(),
        )))

    return results


# ── Dispatcher ────────────────────────────────────────────────────────────────

def parse_bom_file(filepath: str | Path) -> list[dict]:
    """Auto-detect format and return structured BOM rows."""
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    
    # Handle image files
    if ext in {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif'}:
        try:
            from .ocr_processor import parse_bom_image
            return parse_bom_image(filepath)
        except ImportError:
            raise RuntimeError("OCR processor not available. Install: pip install pytesseract pillow")
        except Exception as e:
            raise RuntimeError(f"Image processing failed: {e}")
    
    if ext == ".pdf":
        return parse_pdf(filepath)
    if ext in (".xlsx", ".xls"):
        return parse_xlsx(filepath)
    if ext in (".csv", ".tsv", ".txt"):
        return parse_csv(filepath)
    # Unknown extension — try CSV, then plain text
    try:
        return parse_csv(filepath)
    except Exception:
        text = filepath.read_text(encoding="utf-8", errors="replace")
        return _parse_text(text)
